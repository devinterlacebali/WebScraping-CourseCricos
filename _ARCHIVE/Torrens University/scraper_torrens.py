#!/usr/bin/env python3
"""
Hybrid scraper for Torrens University Australia (03389E).
- Fetches course URLs from sitemap.xml
- Validates & scrapes SSR course pages for title, duration, intake, CRICOS, delivery, location
- Matches CRICOS code against CSV for fee, level, field data
- Generates XLSX + SQL output (only for actual course pages, not category listings)
"""

import csv
import json
import os
import re
import subprocess
import sys
import time
import urllib.parse
from collections import OrderedDict
from datetime import datetime

# === CONFIG ===
PROVIDER_CODE = "03389E"
PROVIDER_NAME = "Torrens University Australia"
BASE_URL = "https://www.torrens.edu.au"
SITEMAP_URL = f"{BASE_URL}/sitemap.xml"
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.join(WORK_DIR, "..")
CSV_PATH = os.path.join(PROJECT_DIR, "cricos-courses.csv")
OUTPUT_NAME = "torrens_scraped"
DELAY = 0.35

CURL_CMD = [
    "curl", "-sS", "-L",
    "-A", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "--connect-timeout", "12", "--max-time", "22",
]


def run_curl(url):
    try:
        r = subprocess.run(CURL_CMD + [url], capture_output=True, text=True, timeout=25)
        return r.stdout
    except Exception:
        return ""


def fetch_sitemap():
    """Download sitemap.xml and extract candidate course URLs."""
    print("📡 Fetching sitemap...")
    xml = run_curl(SITEMAP_URL)
    if not xml:
        print("❌ Failed to fetch sitemap!")
        return []

    urls = re.findall(r'<loc>(.*?)</loc>', xml)
    print(f"  Found {len(urls)} total URLs in sitemap")

    # Filter to paths like /courses/{area}/{slug} (≥3 segments under /courses/)
    candidates = []
    for u in urls:
        path = urllib.parse.urlparse(u).path
        parts = path.strip("/").split("/")
        if parts[0] == "courses" and len(parts) >= 3:
            candidates.append(u)

    return sorted(set(candidates))


def is_course_page(html):
    """Detect if HTML is an actual course detail page (not a category listing)."""
    if not html:
        return False
    # Real course pages have CRICOS code
    if re.search(r'CRICOS\s+\d{6,7}[A-Z]?', html, re.IGNORECASE):
        return True
    # Or structured panels
    panels = re.findall(r'course-card-panel__label[^>]*>(.*?)</div>', html, re.DOTALL)
    non_empty = [p for p in panels if len(re.sub(r'<[^>]+>', '', p).strip()) > 3]
    return len(non_empty) >= 4  # 4+ labeled panels = course page


def parse_course_page(url, html):
    """Parse SSR course page HTML to extract course details."""
    data = OrderedDict()
    data["url"] = url

    # Course name from h1
    h1 = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.DOTALL)
    name = h1.group(1).strip() if h1 else ""
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'<br\s*/?>', ' ', name)
    data["course_name"] = re.sub(r'<[^>]+>', '', name).strip()

    # CRICOS
    m = re.search(r'CRICOS\s+(\d{6,7}[A-Z]?)', html, re.IGNORECASE)
    if not m:
        m = re.search(r'CRICOS[:\s]*(\d{6,7}[A-Z]?)', html, re.IGNORECASE)
    data["cricos_course_code"] = m.group(1).strip() if m else ""

    # Panel values by label
    panel_values = re.findall(r'course-card-panel__value[^>]*>(.*?)</div>', html, re.DOTALL)
    panel_labels = re.findall(r'course-card-panel__label[^>]*>(.*?)</div>', html, re.DOTALL)

    def clean(t):
        t = re.sub(r'<[^>]+>', ' ', t).strip()
        t = re.sub(r'\s+', ' ', t)
        return t

    vals = [clean(p) for p in panel_values if clean(p) and "div{" not in clean(p)[:5]]
    lbls = [clean(l) for l in panel_labels if clean(l) and "div{" not in clean(l)[:5]]

    data["delivery_mode"] = ""
    data["location"] = ""
    data["student_type"] = ""
    data["duration"] = ""
    data["intake_date"] = ""

    for i, label in enumerate(lbls):
        val = vals[i] if i < len(vals) else ""
        lo = label.lower()
        if "delivery" in lo or "mode" in lo:
            data["delivery_mode"] = val
        elif "location" in lo:
            data["location"] = val
        elif "duration" in lo:
            data["duration"] = val
        elif "intake" in lo or "start" in lo or "date" in lo:
            if not data["intake_date"]:
                data["intake_date"] = val
        elif "domestic" in lo or "international" in lo:
            data["student_type"] = val

    # Description
    desc = re.search(r'course-overview__left[^>]*>(.*?)</div>\s*</div>\s*</div>', html, re.DOTALL)
    if desc:
        t = re.sub(r'<[^>]+>', ' ', desc.group(1))
        data["description"] = re.sub(r'\s+', ' ', t).strip()[:2000]
    else:
        data["description"] = ""

    return data


def load_csv_courses():
    print(f"📁 Loading CSV from {CSV_PATH}")
    csv_courses = {}
    try:
        with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                if row.get("CRICOS Provider Code", "").strip() == PROVIDER_CODE:
                    csv_courses[row.get("CRICOS Course Code", "").strip()] = row
        print(f"  Found {len(csv_courses)} Torrens courses in CSV")
    except FileNotFoundError:
        print(f"  ⚠️  CSV not found")
    return csv_courses


def merge_data(page_data, csv_row):
    merged = OrderedDict()
    merged["CRICOS Provider Code"] = PROVIDER_CODE
    merged["Institution Name"] = PROVIDER_NAME
    merged["url"] = page_data.get("url", "")
    merged["course_name"] = page_data.get("course_name", "")
    merged["cricos_course_code"] = page_data.get("cricos_course_code", "")

    if csv_row:
        merged["CRICOS Course Code"] = csv_row.get("CRICOS Course Code", "")
        merged["Course Name (CSV)"] = csv_row.get("Course Name", "")
        merged["Duration (Weeks)"] = csv_row.get("Duration (Weeks)", "")
        merged["Tuition Fee"] = csv_row.get("Tuition Fee", "")
        merged["Non Tuition Fee"] = csv_row.get("Non Tuition Fee", "")
        merged["Estimated Total Course Cost"] = csv_row.get("Estimated Total Course Cost", "")
        merged["Course Level"] = csv_row.get("Course Level", "")
        merged["Field of Education"] = csv_row.get("Field of Education 1 Broad Field", "")
        merged["VET National Code"] = csv_row.get("VET National Code", "")
        merged["Work Component"] = csv_row.get("Work Component", "")
        merged["Expired"] = csv_row.get("Expired", "")
        merged["Course Language"] = csv_row.get("Course Language", "")
    else:
        for k in ["CRICOS Course Code", "Course Name (CSV)", "Duration (Weeks)", "Tuition Fee",
                   "Non Tuition Fee", "Estimated Total Course Cost", "Course Level",
                   "Field of Education", "VET National Code", "Work Component", "Expired", "Course Language"]:
            merged[k] = ""

    merged["duration (page)"] = page_data.get("duration", "")
    merged["delivery_mode"] = page_data.get("delivery_mode", "")
    merged["location"] = page_data.get("location", "")
    merged["student_type"] = page_data.get("student_type", "")
    merged["intake_date"] = page_data.get("intake_date", "")
    merged["description"] = page_data.get("description", "")
    return merged


def esc(val):
    if val is None:
        return "NULL"
    return "'" + str(val).replace("'", "''") + "'"


def gen_sql(row, tbl="courses"):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cols = [
        "cricos_provider_code", "institution_name", "cricos_course_code",
        "course_name", "course_name_csv", "course_url",
        "course_level", "field_of_education",
        "duration", "duration_weeks",
        "delivery_mode", "location", "student_type", "intake_date",
        "tuition_fee", "non_tuition_fee", "estimated_total_cost",
        "course_language", "work_component", "vet_national_code",
        "expired", "description", "created_at",
    ]
    vals = [esc(row.get(c, "")) for c in [
        "CRICOS Provider Code", "Institution Name", "cricos_course_code",
        "course_name", "Course Name (CSV)", "url",
        "Course Level", "Field of Education",
        "duration (page)", "Duration (Weeks)",
        "delivery_mode", "location", "student_type", "intake_date",
        "Tuition Fee", "Non Tuition Fee", "Estimated Total Course Cost",
        "Course Language", "Work Component", "VET National Code",
        "Expired", "description",
    ]] + [esc(now)]
    return f"INSERT INTO {tbl} (\n  " + ",\n  ".join(cols) + "\n) VALUES (\n  " + ",\n  ".join(vals) + "\n);"


def output_xlsx(all_data, filepath):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Torrens Scraped"
        if not all_data:
            print("  ⚠️  No data")
            return False
        headers = list(all_data[0].keys())
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="BF2A45", end_color="BF2A45", fill_type="solid")
        ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bd = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, bd
        for ri, rd in enumerate(all_data, 2):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=ri, column=ci, value=rd.get(h, ""))
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = bd
        for ci in range(1, len(headers) + 1):
            ml = max(len(str(headers[ci - 1])), 10)
            for ri in range(2, min(len(all_data) + 2, 52)):
                ml = max(ml, min(len(str(ws.cell(row=ri, column=ci).value or "")), 60))
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(ml + 3, 80)
        ws.freeze_panes = "A2"
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        wb.save(filepath)
        print(f"  ✅ XLSX saved: {filepath}")
        return True
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=False)
        return output_xlsx(all_data, filepath)


def output_sql(all_data, filepath):
    lines = [
        f"-- Torrens University Australia (03389E) scraped on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"-- Total courses: {len(all_data)}", "",
        "BEGIN TRANSACTION;", "",
    ]
    for row in all_data:
        lines.append(gen_sql(row))
        lines.append("")
    lines.append("COMMIT;")
    lines.append("")
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  ✅ SQL saved: {filepath}")


def main():
    print("=" * 70)
    print(f"  🏫 Torrens University Australia — Hybrid Scraper")
    print(f"  📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    csv_courses = load_csv_courses()
    sitemap_urls = fetch_sitemap()
    if not sitemap_urls:
        print("❌ No candidate URLs found. Exiting.")
        return

    all_data = []
    course_count = 0
    matched_count = 0
    skipped = 0
    errors = []

    print(f"\n📡 Scanning {len(sitemap_urls)} candidate URLs...")
    print("-" * 70)

    for idx, url in enumerate(sitemap_urls, 1):
        slug = url.replace(BASE_URL, "")
        html = run_curl(url)

        if not html:
            errors.append(url)
            print(f"  [{idx:3d}] ⚠️  Empty response: {slug}")
            continue

        if not is_course_page(html):
            skipped += 1
            print(f"  [{idx:3d}] ⛔ Skip (landing page): {slug}")
            if idx < len(sitemap_urls):
                time.sleep(DELAY)
            continue

        page_data = parse_course_page(url, html)
        cricos = page_data.get("cricos_course_code", "")
        csv_row = csv_courses.get(cricos, None)
        merged = merge_data(page_data, csv_row)
        all_data.append(merged)
        course_count += 1
        if csv_row:
            matched_count += 1

        name = page_data.get("course_name", "")
        print(f"  [{idx:3d}] ✅ {slug}")
        print(f"         📖 {name[:80]}")
        print(f"         🔑 CRICOS: {cricos or 'N/A'} | CSV: {'✅' if csv_row else '❌'}")

        if idx < len(sitemap_urls):
            time.sleep(DELAY)

    # Summary
    print("\n" + "=" * 70)
    print(f"  📊 SUMMARY")
    print(f"     Candidates:   {len(sitemap_urls)}")
    print(f"     ✅ Courses:    {course_count}")
    print(f"     ⛔ Skipped:    {skipped} (category/landing pages)")
    print(f"     ✅ CSV match:  {matched_count}")
    print(f"     ❌ Errors:     {len(errors)}")
    print("=" * 70)

    # Save outputs
    print("\n💾 Saving outputs...")
    output_xlsx(all_data, os.path.join(WORK_DIR, f"{OUTPUT_NAME}.xlsx"))
    output_sql(all_data, os.path.join(WORK_DIR, f"{OUTPUT_NAME}.sql"))

    json_path = os.path.join(WORK_DIR, f"{OUTPUT_NAME}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_data, f, indent=2, ensure_ascii=False)
    print(f"  ✅ JSON saved: {json_path}")

    print(f"\n✅ Done! {course_count} courses scraped, {matched_count} matched CSV.")


if __name__ == "__main__":
    main()
