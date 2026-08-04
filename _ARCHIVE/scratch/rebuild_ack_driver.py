"""Rebuild Acknowledge Education driver xlsx from original course pages."""
import requests, re
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os

DIR = r'C:\Users\Dewa(Interlace)\Documents\Interlace Code\WebScraping-CourseCricos\Acknowledge Education'
EXCEL_PATH = os.path.join(DIR, 'acknowledgeeducation.xlsx')

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

# Source pages - we know these courses exist on acknowledgeeducation.edu.au/courses/
course_urls = [
    'https://www.acknowledgeeducation.edu.au/courses/diploma-of-nursing',
    'https://www.acknowledgeeducation.edu.au/courses/certificate-iii-in-patisserie',
    'https://www.acknowledgeeducation.edu.au/courses/advanced-diploma-of-hospitality-management',
    'https://www.acknowledgeeducation.edu.au/courses/diploma-of-hospitality-management',
    'https://www.acknowledgeeducation.edu.au/courses/certificate-iv-in-patisserie',
    'https://www.acknowledgeeducation.edu.au/courses/certificate-iv-in-kitchen-management',
    'https://www.acknowledgeeducation.edu.au/courses/certificate-iii-in-commercial-cookery',
    'https://www.acknowledgeeducation.edu.au/courses/diploma-of-civil-construction-design',
    'https://www.acknowledgeeducation.edu.au/courses/advanced-diploma-of-civil-construction-design',
    'https://www.acknowledgeeducation.edu.au/courses/master-of-social-work-qualifying',
    'https://www.acknowledgeeducation.edu.au/courses/bachelor-of-community-services',
    'https://www.acknowledgeeducation.edu.au/courses/graduate-certificate-of-health-and-community',
    'https://www.acknowledgeeducation.edu.au/courses/bachelor-of-tourism-and-hospitality-management',
    'https://www.acknowledgeeducation.edu.au/courses/bachelor-of-construction-project-management',
    'https://www.acknowledgeeducation.edu.au/courses/master-of-counselling',
    'https://www.acknowledgeeducation.edu.au/courses/bachelor-of-social-work',
    'https://www.acknowledgeeducation.edu.au/courses/graduate-diploma-of-early-childhood-education',
    'https://www.acknowledgeeducation.edu.au/courses/diploma-of-community-services',
    'https://www.acknowledgeeducation.edu.au/courses/ielts-preparation',
    'https://www.acknowledgeeducation.edu.au/courses/english-for-secondary-schools',
    'https://www.acknowledgeeducation.edu.au/courses/vce-year-11-and-12',
    'https://www.acknowledgeeducation.edu.au/courses/english-for-academic-purposes',
    'https://www.acknowledgeeducation.edu.au/courses/general-english',
]

# Try to get from page title + cricos
wb = Workbook()
ws = wb.active
ws.title = "Courses"
ws.append(['cricos', 'title', 'url'])

existing_urls = set()
count = 0

# First, try to extract from the old file (backup from step 1)
old_path = os.path.join(DIR, 'acknowledge_courses_update.sql')
if os.path.exists(old_path):
    with open(old_path, encoding='utf-8') as f:
        sql = f.read()
    for m in re.finditer(r"apply_form = '([^']+)'", sql):
        url = m.group(1)
        if url and url.startswith('http') and url not in existing_urls:
            existing_urls.add(url)

print(f"Found {len(existing_urls)} URLs from SQL backup")

# For any we don't have, fetch from pages
fetched_count = 0
for url in course_urls:
    if url in existing_urls:
        continue
    try:
        r = requests.get(url, headers=headers, timeout=60)
        soup = BeautifulSoup(r.text, 'html.parser')
        title = soup.find('h1')
        title_text = title.get_text(strip=True) if title else url.split('/')[-1]
        # CRICOS
        cricos = ""
        for span in soup.find_all('span'):
            text = span.get_text(strip=True)
            if re.match(r'^[0-9]{6,7}[A-Za-z]?$', text):
                cricos = text
                break
        ws.append([cricos, title_text, url])
        existing_urls.add(url)
        fetched_count += 1
        print(f"  + {title_text[:40]} | {cricos}")
    except Exception as e:
        print(f"  ✗ {url}: {e}")

# Also add from existing_urls that weren't in course_urls
for url in existing_urls:
    try:
        r = requests.get(url, headers=headers, timeout=60)
        soup = BeautifulSoup(r.text, 'html.parser')
        title = soup.find('h1')
        title_text = title.get_text(strip=True) if title else url.split('/')[-1]
        cricos = ""
        for span in soup.find_all('span'):
            text = span.get_text(strip=True)
            if re.match(r'^[0-9]{6,7}[A-Za-z]?$', text):
                cricos = text
                break
        # Check if already in sheet
        already = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == url:
                already = True
                break
        if not already:
            ws.append([cricos, title_text, url])
            print(f"  + (sql) {title_text[:40]} | {cricos}")
    except Exception as e:
        pass

wb.save(EXCEL_PATH)
print(f"\n✅ Saved {ws.max_row - 1} courses to {EXCEL_PATH}")
