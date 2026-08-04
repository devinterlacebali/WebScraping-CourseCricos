import openpyxl

wb = openpyxl.load_workbook("Victoria University/vu.xlsx")
sheet = wb.active

headers = [cell.value for cell in sheet[1]]
print("Headers:", headers)

title_idx = headers.index("title")
url_idx = headers.index("url")
cricos_idx = headers.index("cricos")
fee_idx = headers.index("offshore_tuition_fee")
dur_idx = headers.index("course_duration_per_week")

anomalies = []
for row_idx in range(2, sheet.max_row + 1):
    title = sheet.cell(row=row_idx, column=title_idx+1).value
    url = sheet.cell(row=row_idx, column=url_idx+1).value
    cricos = sheet.cell(row=row_idx, column=cricos_idx+1).value
    fee = sheet.cell(row=row_idx, column=fee_idx+1).value
    dur = sheet.cell(row=row_idx, column=dur_idx+1).value
    
    try:
        # Some values could be strings like "NULL" or have currency symbols
        if fee is None or str(fee).strip().upper() == "NULL" or str(fee).strip() == "":
            fee_val = None
        else:
            fee_val = float(str(fee).replace("$", "").replace(",", "").strip())
            
        if dur is None or str(dur).strip().upper() == "NULL" or str(dur).strip() == "":
            dur_val = None
        else:
            dur_val = float(dur)
    except Exception as e:
        print(f"Row {row_idx}: Error parsing fee '{fee}' or dur '{dur}': {e}")
        continue
        
    if fee_val is not None and dur_val is not None and dur_val > 0:
        annual = fee_val / dur_val * 52
        if annual > 100000 or annual < 5000:
            anomalies.append((row_idx, title, url, cricos, fee_val, dur_val, annual))

print(f"\nFound {len(anomalies)} anomalies in vu.xlsx:")
for r, t, u, c, f, d, a in anomalies:
    print(f"Row {r} | Title: {t} | CRICOS: {c} | Fee: {f} | Dur: {d} | Annual: {a:.2f} | URL: {u}")
