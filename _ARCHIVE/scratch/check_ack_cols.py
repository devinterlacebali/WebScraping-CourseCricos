"""Check actual column structure of Acknowledge Excel."""
from openpyxl import load_workbook
wb = load_workbook('Acknowledge Education/acknowledgeeducation.xlsx')
ws = wb.active
print('Header:', [c.value for c in ws[1]])
print('Row 2:', [c.value for c in ws[2]])
