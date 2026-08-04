import openpyxl

wb = openpyxl.load_workbook("The University Of Adelaide/adelaide.xlsx")
sheet = wb.active

headers = [cell.value for cell in sheet[1]]
url_idx = headers.index("url")
title_idx = headers.index("title")

arts_rows = []
for row_idx in range(2, sheet.max_row + 1):
    url = sheet.cell(row=row_idx, column=url_idx+1).value
    title = sheet.cell(row=row_idx, column=title_idx+1).value
    if "bachelor-of-arts" in str(url):
        arts_rows.append((row_idx, title, url))

print(f"Found {len(arts_rows)} rows matching 'bachelor-of-arts' in adelaide.xlsx:")
for r, t, u in arts_rows:
    print(f"Row {r} | Title: {t} | URL: {u}")
