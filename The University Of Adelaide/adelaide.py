"""
Adelaide University course scraper (requests, plain HTTP).

Extracts data from meta tags + page body of adelaide.edu.au course pages.
Fees are quoted as "indicative annual" on the page; multiply by years for total.
"""
import os
import re
import sys
from urllib.parse import urljoin
import requests
from bs4 import BeautifulSoup

if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# --- constants -------------------------------------------------------------
PROVIDER_CODE = "04249J"                          # Adelaide University (new merged entity)
SLUG = "adelaide"
DIR = "The University Of Adelaide"
EXCEL_PATH = f"{DIR}/{SLUG}.xlsx"
SQL_PATH = f"{DIR}/{SLUG}_courses_update.sql"
REGISTER_CSV = "cricos-courses.csv"

MONTHS = {
    "jan": "January", "feb": "February", "mar": "March", "apr": "April",
    "may": "May", "jun": "June", "jul": "July", "aug": "August",
    "sep": "September", "sept": "September", "oct": "October",
    "nov": "November", "dec": "December", "january": "January",
    "february": "February", "march": "March", "april": "April", "june": "June",
    "july": "July", "august": "August", "september": "September",
    "october": "October", "november": "November", "december": "December",
}
MONTH_ORDER = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)

# --- shared helpers -----------------------------------------------------------
def clean_html(html: str) -> str:
    if not html:
        return ""
    html = re.sub(r"\s+", " ", html)
    return html.replace("'", "''").strip()

def clean_numeric_fee(val) -> str:
    if val is None or str(val).strip().lower() in ("nan", "null", "n/a", "", "none"):
        return "NULL"
    v = re.sub(r"[^\d\.]", "", str(val))
    if not v:
        return "NULL"
    n = float(v)
    return str(int(n)) if n.is_integer() else str(n)

ALLOWED_TAGS = {"p", "ul", "ol", "li", "strong", "b", "em", "i", "a", "br", "h5",
                "table", "thead", "tbody", "tr", "td", "th"}

def sanitise(html: str) -> str:
    if not html:
        return ""
    frag = BeautifulSoup(html, "html.parser")
    for t in frag.find_all(["style", "script", "noscript", "form", "iframe", "img", "svg", "button"]):
        t.decompose()
    for t in frag.find_all(True):
        for a in list(t.attrs):
            if a != "href":
                del t[a]
    for t in frag.find_all("span"):
        t.unwrap()
    while True:
        div = frag.find("div")
        if div is None:
            break
        if div.find(["p", "ul", "ol", "li", "div", "table", "h5"]):
            div.unwrap()
        else:
            div.name = "p"
    for p in frag.find_all("p"):
        s = p.get_text(strip=True)
        if s.endswith(":") and len(s) < 60 and not p.find(["strong", "b", "a"]):
            p.string = ""
            strong = frag.new_tag("strong")
            strong.string = s
            p.append(strong)
    for t in frag.find_all(True):
        if t.name not in ALLOWED_TAGS:
            t.unwrap()
    for t in frag.find_all(["p", "li", "strong", "b", "em", "i"]):
        if not t.get_text(strip=True) and not t.find("br"):
            t.decompose()
    return str(frag)

def txt(el):
    return re.sub(r"\s+", " ", el.get_text()).strip()

def months_in(text):
    found = []
    for tok in re.findall(r"[A-Za-z]{3,9}", str(text)):
        k = tok.lower()
        if k in MONTHS and MONTHS[k] not in found:
            found.append(MONTHS[k])
    return found

def parse_years_to_weeks(text):
    """Convert '4 year(s) full-time' to weeks."""
    m = re.search(r"([\d.]+)\s*years?", str(text), re.I)
    if m:
        try:
            years = float(m.group(1))
            return str(int(round(years * 52)))
        except ValueError:
            pass
    m = re.search(r"([\d.]+)\s*months?", str(text), re.I)
    if m:
        try:
            months = float(m.group(1))
            return str(int(round(months * 4.33)))
        except ValueError:
            pass
    m = re.search(r"(\d+)\s*weeks?", str(text), re.I)
    if m:
        return m.group(1)
    return ""

def get_page(url, tries=3):
    last = RuntimeError("unreachable")
    for i in range(tries):
        try:
            return SESSION.get(url, timeout=60)
        except Exception as e:
            last = e
            import time
            time.sleep(1.5 * (i + 1))
    raise last

# --- extraction ---------------------------------------------------------------
def extract_course_description(soup):
    """Build structured description from sections."""
    desc_parts = []
    
    # Overview section
    overview_heading = soup.find(lambda t: t.name in ('h2','h3','h4','strong') and 'overview' in t.get_text(strip=True).lower())
    if overview_heading:
        content = []
        for tag in overview_heading.find_all_next():
            if tag.name in ('h2','h3','h4') and tag != overview_heading and 'overview' not in tag.get_text(strip=True).lower():
                break
            content.append(str(tag))
        if content:
            desc_parts.append(f"<h4>Overview</h4>{sanitise(''.join(content))}")
    
    # If no overview found, get anything before 'Entry requirements'
    if not desc_parts:
        body = soup.find('main') or soup.find('body')
        if body:
            text = body.get_text()
            idx = text.lower().find('entry requirements')
            if idx > 200:
                before = text[:idx]
                desc_parts.append(f"<h4>Overview</h4><p>{clean_html(before[:2000])}</p>")
    
    return "\n".join(desc_parts)

def extract_entry_requirements(soup):
    """Entry requirements section."""
    parts = []
    
    # Find all requirement-related sections
    for heading_text in ['Entry requirements', 'Admission criteria', 'English language requirements', 'International admissions by country']:
        heading = soup.find(lambda t: t.name in ('h2','h3','h4','strong') and heading_text.lower() in t.get_text(strip=True).lower())
        if heading:
            content = []
            for tag in heading.find_all_next():
                if tag.name in ('h2','h3','h4') and tag != heading:
                    break
                if tag.name in ('p','ul','ol','li','div','table','h5'):
                    content.append(str(tag))
            if content:
                parts.append(f"<h4>{heading_text}</h4>{sanitise(''.join(content))}")
    
    return "\n".join(parts)

def extract_duration(html, soup):
    """Extract duration in weeks from meta tag or body."""
    # From meta tag
    m = re.search(r'timeRequired"\s+content="([^"]+)"', html)
    if m:
        weeks = parse_years_to_weeks(m.group(1))
        if weeks:
            return weeks
    # Fallback: body text
    body = soup.get_text()
    return parse_years_to_weeks(body)

def extract_fees(html, soup, duration_weeks):
    """Extract fees. Annual fee * years = total."""
    years = 1
    m = re.search(r'timeRequired"\s+content="([^"]+)"', html)
    if m:
        ym = re.search(r"([\d.]+)\s*years?", m.group(1), re.I)
        if ym:
            years = float(ym.group(1))
    
    body_text = soup.get_text()
    offshore = "NULL"
    
    # Fee lives in: <div class="degree-details-content-section-subtitle"><span>$54,900</span></div>
    # Find the subtitle that contains a dollar amount
    fee_subtitle = None
    for sub in soup.select('div.degree-details-content-section-subtitle'):
        txt = sub.get_text()
        if '$' in txt:
            fee_subtitle = sub
            break
    if fee_subtitle:
        dm = re.search(r'\$([0-9,]+)', fee_subtitle.get_text())
        if dm:
            annual = float(dm.group(1).replace(',', ''))
            offshore = str(int(round(annual * years)))
    
    # Enrolment fee: $150 application fee
    enrolment = "NULL"
    app_fee_match = re.search(r'application fee (?:of )?AUD?\$([0-9,]+)', body_text, re.I)
    if not app_fee_match:
        app_fee_match = re.search(r'\$([0-9,]+).*?application fee', body_text, re.I)
    if app_fee_match:
        enrolment = clean_numeric_fee(app_fee_match.group(1))
    
    return offshore, "NULL", enrolment, "NULL"

def extract_intake_months(html, soup):
    """Intake months from meta or body."""
    m = re.search(r'startMonth"\s+content="([^"]+)"', html)
    if m:
        return months_in(m.group(1))
    return months_in(soup.get_text())

def extract_cricos(html):
    """CRICOS from meta tag."""
    m = re.search(r'cricosCode"\s+content="([^"]+)"', html)
    return m.group(1) if m else ""

# --- per course ---------------------------------------------------------------
def scrape_course(row):
    cricos = str(row[0]).strip() if row[0] else ""
    title = str(row[1]).strip() if row[1] else ""
    url = str(row[2]).strip() if row[2] else ""
    
    if cricos.lower() in ("nan", "none", "null", ""):
        cricos = ""
    m = re.search(r"[0-9A-Z]{5,8}", cricos)
    cricos = m.group(0) if m else ""
    
    d = {"cricos": cricos, "title": title, "url": url, "course_description": "",
         "course_duration_per_week": "", "offshore_tuition_fee": "NULL",
         "onshore_tuition_fee": "NULL", "enrolment_fee": "NULL", "materials_fee": "NULL",
         "entry_requirements": "", "apply_form": url, "intake_months": []}
    
    try:
        r = get_page(url)
        html = r.text
        soup = BeautifulSoup(html, "html.parser")
        
        d["course_description"] = clean_html(extract_course_description(soup))
        d["entry_requirements"] = clean_html(extract_entry_requirements(soup))
        d["course_duration_per_week"] = extract_duration(html, soup)
        d["offshore_tuition_fee"], d["onshore_tuition_fee"], d["enrolment_fee"], d["materials_fee"] = extract_fees(html, soup, d["course_duration_per_week"])
        d["intake_months"] = extract_intake_months(html, soup)
        
        # CRICOS from xlsx is authoritative, but fill from page if blank
        if not d["cricos"]:
            page_cricos = extract_cricos(html)
            if page_cricos:
                d["cricos"] = page_cricos
        
        print(f"✅ {title[:60]}")
    except Exception as e:
        print(f"❌ {url[:80]}: {e}")
    
    return d

# --- main ---------------------------------------------------------------------
def main():
    import time as ttime
    from openpyxl import load_workbook
    
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel not found: {EXCEL_PATH}")
        return
    
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    results = []
    for i, row in enumerate(rows, 1):
        print(f"[{i}/{len(rows)}] ", end="")
        d = scrape_course(row)
        results.append(d)
        if i % 20 == 0:
            ttime.sleep(2)  # be polite
    
    # Collect intake months
    months = set()
    for d in results:
        months.update(d["intake_months"])
    intake_date = ", ".join(m for m in MONTH_ORDER if m in months)
    
    # Write SQL
    with open(SQL_PATH, "w", encoding="utf-8") as f:
        f.write("-- UPDATE provider institution\n")
        f.write("UPDATE provider_institution SET\n")
        f.write(f"    intake_date = '{intake_date}',\n    updated_at = NOW()\n")
        f.write(f"WHERE cricos_provider_code = '{PROVIDER_CODE}';\n\n")
        
        for d in results:
            if not d["cricos"]:
                f.write(f"-- ⚠️ Skipped (no CRICOS): {d['title']} | {d['url']}\n\n")
                continue
            
            desc = d["course_description"]
            entry = d["entry_requirements"]
            apply = d["apply_form"].replace("'", "''")
            
            f.write(f"UPDATE courses SET\n")
            f.write(f"    course_description = '{desc}',\n")
            f.write(f"    course_duration_per_week = {d['course_duration_per_week'] or 'NULL'},\n")
            f.write(f"    offshore_tuition_fee = {d['offshore_tuition_fee']},\n")
            f.write(f"    onshore_tuition_fee = {d['onshore_tuition_fee']},\n")
            f.write(f"    enrolment_fee = {d['enrolment_fee']},\n")
            f.write(f"    materials_fee = {d['materials_fee']},\n")
            f.write(f"    entry_requirements = '{entry}',\n")
            f.write(f"    apply_form = '{apply}',\n")
            f.write(f"    updated_at = NOW()\n")
            f.write(f"WHERE cricos_course_code = '{d['cricos']}';\n\n")
    
    # Write Excel (rewrite driver with enriched data)
    from openpyxl import Workbook
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Courses"
    ws2.append(["cricos", "title", "url", "course_duration_per_week",
                "offshore_tuition_fee", "onshore_tuition_fee", "enrolment_fee",
                "materials_fee", "intake", "course_description", "entry_requirements"])
    for d in results:
        ws2.append([
            d["cricos"], d["title"], d["url"],
            int(d["course_duration_per_week"]) if str(d["course_duration_per_week"]).isdigit() else "",
            d["offshore_tuition_fee"] if d["offshore_tuition_fee"] != "NULL" else "",
            d["onshore_tuition_fee"] if d["onshore_tuition_fee"] != "NULL" else "",
            d["enrolment_fee"] if d["enrolment_fee"] != "NULL" else "",
            d["materials_fee"] if d["materials_fee"] != "NULL" else "",
            ", ".join(d["intake_months"]),
            d["course_description"].replace("''", "'")[:32000],
            d["entry_requirements"].replace("''", "'")[:32000],
        ])
    wb2.save(EXCEL_PATH)
    
    print(f"\n✅ {len(results)} courses processed.")
    print(f"   Intake: {intake_date}")
    print(f"   SQL   -> {SQL_PATH}")
    print(f"   xlsx  -> {EXCEL_PATH}")

if __name__ == "__main__":
    main()
