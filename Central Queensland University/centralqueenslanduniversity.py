"""Scraper: Central Queensland University (CQU) - 00219C."""
import requests, re, json, html as html_mod, os
from urllib.parse import urljoin
from bs4 import BeautifulSoup

CRICOS_PROVIDER = "00219C"
INSTITUTION = "Central Queensland University"
SLUG = "central-queensland-university"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, f"{SLUG}.xlsx")
SQL_PATH = os.path.join(BASE_DIR, f"{SLUG}_courses_update.sql")

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
TIMEOUT = 30

def sanitise(text):
    """Clean HTML to a minimal semantic subset."""
    if not text:
        return ""
    # Remove script and style
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL|re.IGNORECASE)
    text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL|re.IGNORECASE)
    # Allow only basic tags
    allowed = re.compile(r'</?(?:p|br|ul|ol|li|strong|em|b|i|u|h[1-6]|div|span|a|table|tr|td|th|thead|tbody|caption)[^>]*>', re.I)
    parts = []
    pos = 0
    for m in allowed.finditer(text):
        parts.append(html_mod.escape(text[pos:m.start()]))
        parts.append(m.group())
        pos = m.end()
    parts.append(html_mod.escape(text[pos:]))
    return ''.join(parts)


def extract_next_data(html):
    """Extract __NEXT_DATA__ JSON."""
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except:
            pass
    return None


def fetch_course_urls():
    """Get all course URLs from sitemap."""
    print("Fetching sitemap...")
    r = requests.get('https://www.cqu.edu.au/sitemap.xml', headers=HEADERS, timeout=30)
    urls = re.findall(r'<loc>(https://www\.cqu\.edu\.au/courses/[^<]+)</loc>', r.text)
    print(f"  Found {len(urls)} course URLs")
    return urls


def scrape_course(url):
    """Scrape a single CQU course page."""
    result = {
        "cricos": "",
        "title": "",
        "url": url,
        "duration_weeks": "",
        "offshore_fee": "",
        "onshore_fee": "",
        "enrolment_fee": "",
        "materials_fee": "",
        "description": "",
        "entry_requirements": "",
        "intake": ""
    }

    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        if r.status_code != 200:
            return result
    except:
        return result

    html = r.text
    soup = BeautifulSoup(html, 'html.parser')
    nd = extract_next_data(html)

    # Title
    title_el = soup.find('h1') or soup.find('title')
    if title_el:
        result["title"] = title_el.get_text(strip=True)

    # Description from meta
    desc_meta = soup.find('meta', attrs={'name': 'description'})
    if desc_meta:
        result["description"] = desc_meta.get('content', '')

    # Also try JSON-LD / __NEXT_DATA__
    description_text = ""
    entry_text = ""
    cricos_code = ""
    duration_years = ""
    fee_amount = ""

    if nd:
        props = nd.get('props', {}).get('pageProps', {})
        course = props.get('course', {}) or {}

        # Title
        if course.get('name'):
            result["title"] = course['name']

        # Description from abstract
        abstract = course.get('abstract', '')
        if abstract:
            abstract_clean = re.sub(r'<[^>]+>', '', abstract)
            abstract_clean = html_mod.unescape(abstract_clean).strip()
            result["description"] = abstract_clean

        # Duration
        dur_info = course.get('duration', '')
        if isinstance(dur_info, str):
            duration_years = dur_info
        elif isinstance(dur_info, dict):
            dur_text = dur_info.get('fullTime', '') or dur_info.get('text', '')
            duration_years = dur_text

        if not duration_years:
            # Try props level
            dur = props.get('duration', '')
            if dur:
                duration_years = dur

        # Entry requirements
        entry = course.get('entryRequirements', '') or props.get('entryRequirements', '')
        if entry:
            entry_text = re.sub(r'<[^>]+>', '', entry)
            entry_text = html_mod.unescape(entry_text).strip()[:500]

        # CRICOS from course data
        cricos_list = course.get('cricosCodes', [])
        if cricos_list:
            cricos_code = cricos_list[0].get('cricos_code', '')
            if not cricos_code and isinstance(cricos_list[0], dict):
                cricos_code = cricos_list[0].get('cricos_code', '')

        # Fee - from factbox or price info
        fee_info = course.get('feeInformation', '') or course.get('fees', '')
        if isinstance(fee_info, str):
            for m in re.finditer(r'\$[0-9,]+\s*(?:per|/)?\s*(year|annum|annual)', fee_info, re.I):
                fee_amount = m.group()

        # Intake
        next_intake = course.get('nextIntake', '') or props.get('nextIntake', '')
        if isinstance(next_intake, dict):
            result["intake"] = next_intake.get('label', '') or next_intake.get('text', '')

        # Also check availabilities
        avail = course.get('availabilities', [])
        if avail and not result["intake"]:
            intakes = set()
            for a in avail:
                term = a.get('term', '')
                year = a.get('year', '')
                if term and year:
                    intakes.add(f"{term} {year}")
            if intakes:
                result["intake"] = ", ".join(sorted(intakes))

    # Fallback: scrape CRICOS from raw HTML
    if not cricos_code:
        for m in re.finditer(r'CRICOS[:\s]*(\d{6,7}[A-Za-z]?)', html):
            cricos_code = m.group(1)
            break

    # Fallback duration from <meta name="courseDuration">
    if not duration_years:
        dur_meta = soup.find('meta', attrs={'name': 'courseDuration'})
        if dur_meta:
            duration_years = dur_meta.get('content', '')

    result["cricos"] = cricos_code

    # Parse duration
    if duration_years:
        m = re.search(r'(\d+\.?\d*)', str(duration_years))
        if m:
            years = float(m.group(1))
            result["duration_weeks"] = str(int(years * 52))

    # Fee - try to extract from page
    # CQU shows "First-year fee" for domestic, need international fee
    # Look for international fee in factboxes
    if not fee_amount:
        for m in re.finditer(r'\$[0-9,]+\s*(?:per\s*year|annual|indicative)', html, re.I):
            ctx = html[max(0, m.start()-50):m.end()+50]
            if 'international' in ctx.lower() or 'visa' in ctx.lower():
                fee_amount = m.group(1) if m.groups() else m.group()
                break

    if fee_amount:
        fee_num = re.sub(r'[^0-9]', '', fee_amount)
        if fee_num:
            result["offshore_fee"] = fee_num

    # Entry requirements from HTML if not from __NEXT_DATA__
    if not entry_text:
        entry_section = soup.find('h2', string=re.compile(r'entry requirement|admission|entry', re.I))
        if entry_section:
            parent = entry_section.find_parent(['div', 'section'])
            if parent:
                entry_text = parent.get_text(strip=True)[:500]

    result["entry_requirements"] = entry_text

    # Intake from HTML if not found
    if not result["intake"]:
        for el in soup.find_all(['span', 'div', 'p']):
            txt = el.get_text(strip=True)
            m = re.search(r'(?:Next\s+)?(?:start\s+)?(term|intake)[:\s]+(.*?)(?:$|\.)', txt, re.I)
            if m:
                result["intake"] = m.group(2).strip()
                break

    return result


def build_driver_xlsx(urls):
    """Build driver xlsx from course URLs."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["cricos", "title", "url", "duration_weeks", "offshore_fee",
               "onshore_fee", "enrolment_fee", "materials_fee",
               "description", "entry_requirements", "intake"])

    for i, url in enumerate(urls, 1):
        data = scrape_course(url)
        ws.append([data["cricos"], data["title"], data["url"],
                   data["duration_weeks"], data["offshore_fee"],
                   data["onshore_fee"], data["enrolment_fee"],
                   data["materials_fee"],
                   sanitise(data["description"]),
                   sanitise(data["entry_requirements"]),
                   data["intake"]])
        if i % 50 == 0:
            print(f"  Scraped {i}/{len(urls)}...")

    wb.save(EXCEL_PATH)
    print(f"  Saved {len(urls)} courses to {EXCEL_PATH}")
    return wb


def generate_sql(ws):
    """Generate SQL UPDATE statements."""
    lines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cricos, title, url, dur, offshore, onshore, enrol, materials, desc, entry, intake = row
        if not cricos:
            continue
        name_esc = str(title or '').replace("'", "''")
        dur_esc = str(dur or 'NULL').replace("'", "''")
        off_esc = str(offshore or 'NULL').replace("'", "''")
        on_esc = str(onshore or 'NULL').replace("'", "''")
        enrol_esc = str(enrol or 'NULL').replace("'", "''")
        mat_esc = str(materials or 'NULL').replace("'", "''")
        desc_esc = str(desc or '').replace("'", "''")
        entry_esc = str(entry or '').replace("'", "''")
        intake_esc = str(intake or '').replace("'", "''")

        sql = f"""UPDATE courses SET
    course_name = '{name_esc}',
    course_url = '{str(url or '').replace("'", "''")}',
    course_duration_per_week = '{dur_esc}',
    offshore_fee = '{off_esc}',
    onshore_fee = '{on_esc}',
    enrolment_fee = '{enrol_esc}',
    materials_fee = '{mat_esc}',
    description = '{desc_esc}',
    entry_requirements = '{entry_esc}',
    intake = '{intake_esc}'
WHERE cricos_course_code = '{str(cricos).replace("'", "''")}';"""
        lines.append(sql)
    return '\n\n'.join(lines)


def main():
    print(f"=== {INSTITUTION} ({CRICOS_PROVIDER}) ===")

    urls = fetch_course_urls()
    if not urls:
        print("No course URLs found!")
        return

    print(f"\nBuilding driver xlsx for {len(urls)} courses...")
    wb = build_driver_xlsx(urls)

    ws = wb.active
    print(f"\nGenerating SQL...")
    sql = generate_sql(ws)
    with open(SQL_PATH, 'w', encoding='utf-8') as f:
        f.write(sql)
    print(f"  SQL saved: {SQL_PATH}")

    # Summary
    cricos_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[0])
    fee_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[4] and str(r[4]).strip())
    print(f"\n✅ Done! {ws.max_row - 1} courses, {cricos_count} with CRICOS, {fee_count} with fee")


if __name__ == '__main__':
    main()
